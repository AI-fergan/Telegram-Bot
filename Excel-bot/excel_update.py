from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import sys
import log
import config

# File name for storing data
data_file = config.EXCEL_FILE
shops_lst = []
workbook = None

def create_excel_file() -> None:
    """
    This function create new data.xlsx file, 
    and set the global workbook by that file.
    """
    global workbook, data_file

    # Create new `Profits` workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Profits"

    # Set the shops title
    sheet["A1"] = "Shop Name"
    workbook.save(data_file)


def update_shop_profit(shop_name: str, profit: float) -> None:
    """
    This function update shop or create new shop and add its profit for today.
    shop_name - the shop name to create \ update.
    profit - shop daily profit.
    """
    global workbook, data_file

    # Get the `Profits` sheet
    sheet = workbook["Profits"]

    # Get todays date
    today = datetime.now().strftime("%Y-%m-%d")

    # Check if todays date already exists as a column
    if today not in [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]:
        # Add today's date as a new column
        sheet.cell(row=1, column=sheet.max_column + 1).value = today

    # Find the column index for todays date
    today_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == today:
            today_col = col
            break

    # Find the shop row or add a new row for the shop
    shop_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == shop_name:
            shop_row = row
            break

    if shop_row is None:  # Add a new row for the shop
        shop_row = sheet.max_row + 1
        sheet.cell(row=shop_row, column=1).value = shop_name

    # Update the profit for the shop in todays column
    sheet.cell(row=shop_row, column=today_col).value = profit

    # Save the workbook
    workbook.save(data_file)

def init_file() -> None:
    """
    This function create new `data.xlsx` file if not exists.
    """
    global data_file, workbook

    # Check if the data file already exists
    if os.path.exists(data_file):
        workbook = load_workbook(data_file)
    else:
        create_excel_file()

def load_shops() -> None:
    """
    This function load all the shops names to the global shops names list.
    """
    global workbook, shops_lst

    # Get `Profits` sheet from the data file
    sheet = workbook["Profits"]

    # Get all the shops names to the global shops list
    for row in range(2, sheet.max_row + 1):
        shops_lst.append(sheet.cell(row=row, column=1).value)

def run(command: list) -> None:
    """
    This function get command params and type,
    then it will run it or do exit(1) if there is an error during the function run. 
    (commonly by missing or wrong params).
    """
    global shops_lst
    
    log.log(0, "Init data file...")
    init_file()

    log.log(0, "Load shops...")
    load_shops()

    # Check if there is an command
    if len(command) == 0:
        log.log(1, "Command")

    # New command handler
    elif command[0] == "new":
        if len(command) != 3: # Need only two params
            log.log(2, "new")
        if command[1] in shops_lst: # Check if the shop name isnt already exists
            log.log(3, "The shop " + command[1])
        shops_lst += command[1] # Add shop name to the shops list
        update_shop_profit(command[1], command[2]) # Update shop profit

    # Add command handler
    elif command[0] == "add":
        if len(command) != 3: # Need only two params
            log.log(2, "add")
        if command[1] not in shops_lst: # Check if the shop name is exists in shop list
            log.log(1, "Shop name")
        update_shop_profit(command[1], command[2]) # Update shop profit

